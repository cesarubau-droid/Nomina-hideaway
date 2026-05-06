# ============================================================
# GENERADOR EXCEL v2.2
# ============================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn


def generar_excel(df: pd.DataFrame, output_path: str = 'output_nomina.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Nomina"

    C_HDR   = '1F4E79'
    C_NOC   = 'D9E1F2'
    C_FER   = 'FFE0E0'
    C_NOSAL = 'FFF2CC'
    C_QBR   = 'E2D9F3'
    C_CONF  = 'EAF0FB'
    C_EXTRA = 'D9F0E2'
    C_LATE  = 'FFCCCC'
    C_OK    = 'FFFFFF'

    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center')
    lft  = Alignment(horizontal='left',   vertical='center')

    def sc(ws, r, c, val, bold=False, color='000000', fillc=None, align=None, sz=10):
        cl = ws.cell(row=r, column=c, value=val)
        cl.font = Font(name='Arial', bold=bold, color=color, size=sz)
        if fillc:
            cl.fill = PatternFill('solid', start_color=fillc)
        cl.alignment = align or ctr
        cl.border = brd
        return cl

    NCOLS = 19
    headers = [
        'ID', 'Nombre', 'Apellido', 'Departamento', 'Tipo',
        'Fecha', 'Feriado',
        'Entrada Real', 'Entrada Redond',
        'Salida Real',  'Salida Redond',
        'Diurnas Ord', 'Mixtas Ord', 'Nocturnas Ord',
        'Extra Diurnas', 'Extra Mixtas', 'Extra Nocturnas',
        'Estado', 'Notas'
    ]
    for ci, h in enumerate(headers, 1):
        sc(ws, 1, ci, h, bold=True, color='FFFFFF', fillc=C_HDR)
    ws.row_dimensions[1].height = 22

    NUM_COLS = {12, 13, 14, 15, 16, 17}

    def row_color(row):
        status = str(row.get('Estado', ''))
        if row.get('Feriado'):          return C_FER
        if 'ajusta' in status:          return C_NOSAL
        if 'Con Acuerdo' in status:     return C_NOC
        if 'Nocturno' in status:        return C_NOC
        if 'Quebrado' in status:        return C_QBR
        if 'Confianza' in status:       return C_CONF
        if '+Extra' in status:          return C_EXTRA
        if 'Tardío' in status:          return C_LATE
        return C_OK

    df_cols = [
        'ID', 'Nombre', 'Apellido', 'Departamento', 'Tipo',
        'Fecha', 'Feriado',
        'Entrada Real', 'Entrada Redond',
        'Salida Real',  'Salida Redond',
        'Diurnas Ord', 'Mixtas Ord', 'Nocturnas Ord',
        'Extra Diurnas', 'Extra Mixtas', 'Extra Nocturnas',
        'Estado', 'Notas'
    ]

    for ri, row in enumerate(df.to_dict('records')):
        er  = ri + 2
        bg  = row_color(row)
        is_fer = bool(row.get('Feriado'))
        txt = 'C00000' if is_fer else '000000'
        for ci, col in enumerate(df_cols, 1):
            val = row.get(col, '')
            sc(ws, er, ci, val, bold=is_fer, color=txt, fillc=bg)
        ws.row_dimensions[er].height = 16

    last_data_row = len(df) + 1
    totals_row    = last_data_row + 1  # fila de totales = una más

    # Escribir fila de totales manualmente con SUBTOTAL
    for ci in range(1, NCOLS + 1):
        col_letter = get_column_letter(ci)
        if ci in NUM_COLS:
            # SUBTOTAL(109, ...) = SUM ignorando filas ocultas por filtro
            formula = f"=SUBTOTAL(109,{col_letter}2:{col_letter}{last_data_row})"
            ws.cell(row=totals_row, column=ci, value=formula)
            ws.cell(row=totals_row, column=ci).font = Font(name='Arial', bold=True)
            ws.cell(row=totals_row, column=ci).fill = PatternFill('solid', start_color='FFE699')
            ws.cell(row=totals_row, column=ci).border = brd
            ws.cell(row=totals_row, column=ci).alignment = ctr
        elif ci == 2:
            ws.cell(row=totals_row, column=ci, value='TOTAL')
            ws.cell(row=totals_row, column=ci).font = Font(name='Arial', bold=True)
            ws.cell(row=totals_row, column=ci).fill = PatternFill('solid', start_color='FFE699')
            ws.cell(row=totals_row, column=ci).border = brd
            ws.cell(row=totals_row, column=ci).alignment = lft
        else:
            ws.cell(row=totals_row, column=ci).fill = PatternFill('solid', start_color='FFE699')
            ws.cell(row=totals_row, column=ci).border = brd

    ws.row_dimensions[totals_row].height = 20

    # ── TABLA — solo hasta last_data_row (sin incluir la fila de totales) ──
    table_ref = f"A1:{get_column_letter(NCOLS)}{last_data_row}"
    table = Table(displayName="Nomina", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False
    )
    table.tableStyleInfo = style

    for i, name in enumerate(headers):
        tc = TableColumn(id=i+1, name=name)
        table.tableColumns.append(tc)

    ws.add_table(table)

    # Anchos
    widths = [6,14,16,20,10,12,12,12,14,12,14,12,11,13,13,11,14,20,40]
    for i, w in enumerate(widths[:NCOLS], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"Excel guardado: {output_path}")
